from excel_common.excel_operation import (
    ExcelParser,
    AbstractExcelService,
    SheetParserSercice,
)
from enum import Enum
from excel_common.item import Item
from pms_dbmodel.project import CustomerData, ProjectData, ProjectService as PS
import openpyxl
from .util import *
from pms_dbmodel.customer import CustomerService
from pms_dbmodel.platform import PlatformService


class SHEETNAME(Enum):
    Home_Project = 0
    L1_Project = 1
    L2_Project = 2
    L3_Project = 3


class PROJECT(Enum):
    NAME = 0
    PLATFROM = 1
    HOME = 2
    ODM = 3
    T2 = 4
    T1 = 5
    OEM = 6
    PRIORITY = 7
    TYPE = 8
    CPM = 9
    AM = 10

    def isCustomerInfo(self):
        return 2 <= self.value and self.value <= 6


class ProjectParserService:
    @classmethod
    def getAllItem(cls, fileName) -> list[Item]:
        result = []

        excel = openpyxl.load_workbook(fileName)
        result += cls.parseHomeProject(fileName, excel)
        result += cls.parseL1Project(fileName, excel)
        result += cls.parseL2Project(fileName, excel)
        result += cls.parseL3Project(fileName, excel)
        return result

    @classmethod
    def parseHomeProject(cls, fileName, excel) -> list[Item]:
        sheet = excel[SHEETNAME.Home_Project.name]

        keyMap = {
            PROJECT.NAME: "Project Name",
            PROJECT.PLATFROM: "Chipset",
            PROJECT.HOME: "Customer",
            PROJECT.PRIORITY: "Business Priority",
            PROJECT.TYPE: "Product Type",
        }

        return SheetParserSercice.parseSheet(fileName, sheet, keyMap, PROJECT)

    @classmethod
    def parseL1Project(cls, fileName, excel) -> list[Item]:
        sheet = excel[SHEETNAME.L1_Project.name]
        keyMap = {
            PROJECT.NAME: "Project Name",
            PROJECT.PLATFROM: "Chipset",
            PROJECT.OEM: "OEM",
            PROJECT.PRIORITY: "Business Priority",
            PROJECT.TYPE: "Product Type",
        }
        return SheetParserSercice.parseSheet(fileName, sheet, keyMap, PROJECT)

    @classmethod
    def parseL2Project(cls, fileName, excel) -> list[Item]:
        sheet = excel[SHEETNAME.L2_Project.name]
        keyMap = {
            PROJECT.NAME: "Project Name",
            PROJECT.PLATFROM: "Chipset",
            PROJECT.OEM: "OEM/Brand",
            PROJECT.ODM: "ODM",
            PROJECT.PRIORITY: "Business Priority",
            PROJECT.TYPE: "Product Type",
        }
        return SheetParserSercice.parseSheet(fileName, sheet, keyMap, PROJECT)

    @classmethod
    def parseL3Project(cls, fileName, excel) -> list[Item]:
        sheet = excel[SHEETNAME.L3_Project.name]
        keyMap = {
            PROJECT.NAME: "Project name",
            PROJECT.PLATFROM: "IC",
            PROJECT.OEM: "OEM",
            PROJECT.T1: "Tier-1",
            PROJECT.T2: "客戶",
            PROJECT.PRIORITY: "Business Priority",
            PROJECT.TYPE: "Product Type",
        }
        return SheetParserSercice.parseSheet(fileName, sheet, keyMap, PROJECT)

    @classmethod
    def getAllPlatforms(cls, data: list[Item]) -> list[str]:
        result = []

        for item in data:
            platform = item.getKey(PROJECT.PLATFROM)
            if platform not in result:
                result.append(platform)
        return result

    @classmethod
    def getAllCustomers(cls, data: list[Item]) -> list[str]:
        result = []

        for item in data:
            for e in PROJECT:
                if e.isCustomerInfo():
                    customer = item.getKey(e)
                    if customer != None and customer not in result:
                        result.append(customer)

        return result

    @classmethod
    def createProjectData(
        cls, data: list[Item], customerArea: dict
    ) -> list[ProjectData]:
        result = []
        projectName = []
        for d in data:
            name = d.getKey(PROJECT.NAME)
            platform = d.getKey(PROJECT.PLATFROM)
            p = ProjectData([platform, name, None])
            cls._addCustomer(d, p, customerArea)

            if name not in projectName:
                result.append(p)
                projectName.append(name)
        return result

    @classmethod
    def _addCustomer(
        cls, data: Item, project: ProjectData, customerArea: dict
    ) -> CustomerData:
        """
        Note: we do not know why we should have MAIN_PLATFROM and MAIN_CUSTOMER in the PROJECT_DATA
        However because the purpose is for insertion, we can ignore this two attribute
        """
        for e in PROJECT:
            if e.isCustomerInfo():
                customer = data.getKey(e)

                if customer == None:
                    continue

                aList = customerArea.get(customer, None)

                if aList == None:
                    raise Exception(
                        "We can not find the customer %s" % customer,
                    )
                if len(aList) > 1 or len(aList) == 0:
                    raise Exception(
                        " customer %s  has more area in the database" % customer,
                    )

                project.addCustomer(CustomerData([aList[0], customer, e.name]))


class ProjectService(AbstractExcelService):
    @classmethod
    def parse(cls, fileName):
        # 1. parsing excel to items
        data = ProjectParserService.getAllItem(fileName)
        # 2. Examine there are some customers or platforms does not exit in database
        cls.examine(data)
        # 3. sort out all items into ProjectData
        customerArea = CustomerService.getCustomerAreaMap()
        result = ProjectParserService.createProjectData(data, customerArea)
        # 4. add or update project
        for p in result:
            PS.addProject(p)

        pass

    @classmethod
    def examine(cls, data: list[Item]) -> dict[str, str]:
        # 1. Check Customer
        customers = CustomerService.getCustomers()
        for c in ProjectParserService.getAllCustomers(data):
            if c not in customers:
                raise Exception(
                    " Can not find customer(%s) in database, please add it " % c
                )

        # 2. Check Platforms
        platforms = PlatformService.getPlatforms()
        for p in ProjectParserService.getAllPlatforms(data):
            if p not in platforms:
                raise Exception(
                    " Can not find platform(%s) in database, please add it " % p
                )

    @classmethod
    def collectIntoOneExcel(cls, homeFileName, l1FileName, l2FileName, l3FileName):
        """
            Collect from several location and output the excel which system can use
        Args:
            homeFileName (_type_): _description_
            l1FileName (_type_): _description_
            l2FileName (_type_): _description_
            l3FileName (_type_): _description_
        """

        home = openpyxl.load_workbook(homeFileName)
        l1 = openpyxl.load_workbook(l1FileName)
        l2 = openpyxl.load_workbook(l2FileName)
        l3 = openpyxl.load_workbook(l3FileName)
        result = openpyxl.Workbook()

        result.save(OUTPUT_FILE)
