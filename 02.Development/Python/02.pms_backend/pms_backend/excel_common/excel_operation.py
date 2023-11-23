from asyncio.log import logger
import logging
import string
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from enum import Enum
from .item import Item


class KEYINFO(Enum):
    Id = 0
    Description = 1
    Status = 2
    ChapterId = 3
    Chapter = 4
    SectionId = 5
    Section = 6
    DocLocation = 7
    Priority = 8
    Note = 9
    Title = 10


NOCOLUMNINX = -1


class OperatorItem(Item):
    def __init__(self, fileName, productType: list):
        super().__init__(fileName, KEYINFO)

        self.type = productType.copy()
        self.priority = []

    def __str__(self):
        return self.id

    def setKeyList(self, row, keyMap: dict, typeMap: dict):
        # set key
        keyInfo = self.getKetInfo()
        for k in keyInfo:
            idx = keyMap[k]
            if idx == -1:
                self.keyValueList.append(None)
            else:
                self.keyValueList.append(row[idx].value)
        # add priority
        for t in self.type:
            idx = typeMap[t]
            self.addPriority(row[idx].value)

    def addPriority(self, priority):
        self.priority.append(priority)

    def addType(self, t):
        self.type.append(t)

    def getPriority(self):
        return self.getKey(KEYINFO.Priority)

    def getId(self):
        return self.getKey(KEYINFO.Id)

    def getChapterId(self):
        return self.getKey(KEYINFO.ChapterId)

    def getChapter(self):
        return self.getKey(KEYINFO.Chapter)


# End Class Items

"""
   class ExcelParser
"""


class ExcelParser:
    logger = logging.getLogger("ExcelParser")

    def __init__(
        self,
        keyRow=0,
        originalKeyMap={},
        keyInfo=KEYINFO,
        typeRow=0,
        originalProductTypeList=[],
    ):
        self.initEmpty()
        self.initialize(keyRow, originalKeyMap, typeRow, originalProductTypeList)
        self.setKeyInfo(keyInfo)

    """
       Initialize method
    """

    def initEmpty(self):
        self.keyMap = {}
        self.typeMap = {}

    def initialize(
        self, keyRow: int, originalKeyMap: dict, typeRow, originalProductTypeList
    ):
        self._originalKeyMap = originalKeyMap.copy()
        self.keyRow = keyRow
        self._originalProductTypeList = originalProductTypeList.copy()
        self.typeRow = typeRow

    def setKeyInfo(self, enum: Enum):
        self._keyInfo = enum

    def getKeyInfo(self) -> Enum:
        return self._keyInfo

    # END Initialize method

    def createMap(self, sheet: Worksheet):
        self.setKeyMap(sheet, self.keyRow, self._originalKeyMap)
        self.setTypeMap(sheet, self.typeRow, self._originalProductTypeList)

    """
      the input of These Group function is <KeyInfo, ExcelColumn> and it will create < KeyInfo, ExcelIndex>
    """

    def setKeyMap(self, sheet, row, keyMap: dict):
        cellValueList = []
        keyInfo = self.getKeyInfo()
        for k in keyInfo:
            val = keyMap.get(k)

            if val == None or val == "":
                print("Please note the key = " + k.name + " is NULL or empty string.")
                val = NOCOLUMNINX
            cellValueList.append(val)
        index = self.__getColumnsIdxBasedOnList(sheet, row, cellValueList)

        self.__setMap(index, keyInfo, self.keyMap)

    def setTypeMap(self, sheet, row, productTypeList: list):
        index = self.__getColumnsIdxBasedOnList(sheet, row, productTypeList)
        self.__setMap(index, productTypeList, self.typeMap)

    def __setMap(self, index: list, cellValueList, dataMap: map):
        i = 0
        for key in cellValueList:
            dataMap[key] = index[i]
            i = i + 1

    def __getColumnsIdxBasedOnList(self, sheet, row, cellValueList: list) -> list:
        if isinstance(cellValueList, list) != True:
            raise Exception("Note that cellValueList is a list")

        result = cellValueList.copy()

        idxMap = ExcelParser.__getIdxMap(sheet, row, cellValueList)

        for i in range(0, len(cellValueList)):
            result[i] = idxMap[cellValueList[i]]
        return result

    @staticmethod
    def __getIdxMap(sheet: Worksheet, row, cellValueList: list) -> dict:
        idxMap = {}
        for v in cellValueList:
            if idxMap.get(v) == None:
                idxMap[v] = -1

        for col in sheet.iter_cols(min_row=row, max_row=row):
            cell = col[0]
            if cell.value:
                index = idxMap.get(cell.value)
                if index != None and index == -1:
                    idxMap[cell.value] = cell.col_idx - 1
        return idxMap

    # End createMap

    def getKeyIndex(self, key):
        return self.keyMap[key]

    def getProductTypeList(self):
        return self._originalProductTypeList

    """
        The List of keyColumn should follow the sequence: status, capability, id, desc, section.
        The Priority value depends on the List of productColumn
    """

    def parse(self, fileName: string, sheetName: string, dataRow):
        excel = openpyxl.load_workbook(fileName)
        sheet = excel[sheetName]

        logger.info(
            "[ExcelParser.parse][BEGIN]: "
            + fileName
            + " MAX ROW = "
            + str(sheet.max_row)
        )
        result = self.parseExcel(fileName, dataRow, sheet)
        logger.info("[ExcelParser.parse][END]: " + fileName)

        excel.close()
        return result

    def parseExcel(self, fileName, dataRow, sheet: Worksheet) -> list[OperatorItem]:
        result = []
        logger.info("[ExcelParser.parseExcel][BEGIN]: MAX ROW = " + str(sheet.max_row))
        idx = 0
        keyInfo = self.getKeyInfo()
        for row in sheet.iter_rows(min_row=dataRow):
            item = OperatorItem(fileName, self.getProductTypeList())

            item.setKeyInfo(keyInfo)
            item.setKeyList(row, self.keyMap, self.typeMap)
            result.append(item)

            idx = idx + 1
            if idx % 500 == 0:
                print(" processing: " + str(idx))

        logger.info("[ExcelParser.parseExcel][END]:")
        return result


# END class ExcelParser
