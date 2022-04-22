import os
import re
import openpyxl
from common import file_operation, util
from common import excel_operation

"""
   getMTRFromExcel
"""
def getMTRFromExcel(fileName, sheetName):
    
    keyMap = {
        excel_operation.KEYINFO.Id: "Global ID", 
        excel_operation.KEYINFO.Description:  "Description", 
        excel_operation.KEYINFO.Status: "Status", 
        excel_operation.KEYINFO.ChapterId:"Heading", 
        excel_operation.KEYINFO.Chapter: "Name", 
        excel_operation.KEYINFO.SectionId: "Heading", 
        excel_operation.KEYINFO.Section: "Name",
        excel_operation.KEYINFO.DocLocation:"TRD",
        excel_operation.KEYINFO.Priority: "Priority",
        excel_operation.KEYINFO.Note:"OEM Compliance"
    }
    productType = [ 
                    "      High-Tier Smartphone"      ,
                    "      Mid-Tier Smartphone"       ,
                    "      Value Smartphone"          ,
                    "      Feature Phone"             ,
                    "      Tablet/Laptop"                    ,
                    "      Mobile Hotspot/Router"     ,
                    "      Fixed Broadband: Router"   ,
                    "      Wearable"                  ,
                    "      Internet-of-Things"        ]
    
    
    excel = openpyxl.load_workbook(fileName)
    sheet = excel[sheetName]
    parser = excel_operation.ExcelParser(2, keyMap, 2, productType)
    parser.createMap(sheet)
    
    result = parser.parseExcel( fileName, 3, sheet)
    
    excel.close()
    return result

# End getMTRFromExcel
"""
   getCRItemFromExcel
"""
def getCRItemFromExcel(fileName, sheetName):
    keyMap = {
        excel_operation.KEYINFO.Id:"Global ID",
        excel_operation.KEYINFO.Status: "Change Type",
        excel_operation.KEYINFO.Description: "Change Reason"
    }
    excel = openpyxl.load_workbook(fileName)
    sheet = excel[sheetName]
    parser = excel_operation.ExcelParser(2, keyMap)
    parser.createMap(sheet)
    
    items = parser.parseExcel( fileName, 3, sheet)
    
    excel.close()
    
    result = {}
    for item in items:
        id = item.getId()
        result[id] = item
    
    return result
#End getCRItemFromExcel   
"""
    We need to sort out the item
    1. Filter out those items that are below the row: "The Following RequirementsXXXXX"
    1. assign status based on CR item
    2. Return a Map<FLD Id, List<Item>>
    3. we need to rearrange correct Chapter Id and Chapter for each item
"""
def sortOut( data: list[excel_operation.Item], crItem:dict):
    #1. assign status based on CR item
    items = []
    r = 2
    for item in data:
        id = item.getId()
        cr = crItem.get(id)
        if cr != None:
            status = cr.getKeyInfo(excel_operation.KEYINFO.Status)
            item.setKey(excel_operation.KEYINFO.Status, status)
        if item.getKeyInfo(excel_operation.KEYINFO.DocLocation) == "The Following Requirements were excluded for this product - do not change anything below this line":
            continue
        items.append(item)
    print("Total ITEM = " + str( len(items) ) )
    
    #2. Prepare a Map
    result = {} 
    key = ""
    chapterId = ""
    chapter = ""
    l = []
    for item in items:
        id = item.getId()
        priority = item.getPriority()
        if priority == "Heading":
            key = id
            chapterId = item.getChapterId()
            chapter = item.getChapter()
        
            l = result.get(key)
            if( l == None):
                l = []
                result[key] = l
        else: l.append(item)
        if key == "GID-FLD-5128":
            i = 0
        # 3. we need to rearrange correct Chapter Id and Chapter for each item
        if priority != "Heading" and (chapterId != "" or chapter != ""):
            item.setKey(excel_operation.KEYINFO.ChapterId, chapterId)
            item.setKey(excel_operation.KEYINFO.Chapter, chapter)
    
    return result
  
  
def outputExcel( fileName, itemMap:dict):
    wb  = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1).value = "FLD_ID"
    sheet.cell(row=1, column=2).value = "MTR_Id"
    sheet.cell(row=1, column=3).value = "Status"
    
    r = 2
    for key in itemMap.keys():
        items :list[excel_operation.Item] = itemMap.get(key) 
        for item in items:
            sheet.cell(row= r, column=1).value = key
            sheet.cell(row= r, column=2).value = item.getId()
            sheet.cell(row= r, column=3).value = item.getKeyInfo(excel_operation.KEYINFO.Note)
            r = r + 1
    
    wb.save(fileName)
             
# End sortOut
def tmo_parser():
    tmp_parser("2020_Q4")

DEFAULT_FOLDER = "D:\\02.Operator\\01.TMO\\"
def tmp_parser( version ) :
    
    print("[tmp_parser][BEGIN] " + version)
    
    #0. Paramters
    folder = DEFAULT_FOLDER + version +"\\"
    prefix = "T-MobileUS_MTR"
    #1. Get fileName in the folder
    fileName = file_operation.getFileNameList( folder, prefix )
    crFileName = ""
    mtrFileName = ""
    for f in fileName:
        if re.search("cr", f):
            crFileName = f
        else: mtrFileName = f
    
    #2. Get each items in the files
    sheetName = "MTR"
    
    #2.1 Get Critical Column
    item = getMTRFromExcel(mtrFileName, sheetName)
    crItem = getCRItemFromExcel(crFileName, sheetName)
    itemMap = sortOut(item, crItem)
    
    if os.path.isdir(util.OUTPUT_FOLDER) != True:
        os.makedirs(util.OUTPUT_FOLDER)     
    outputFileName = util.OUTPUT_FOLDER+"tmo_"+version+".xlsx"
    
    outputExcel(outputFileName, itemMap)
    
    print("[tmp_parser][END] " + version)
if __name__ == '__main__':
    tmo_parser()