
import openpyxl
from common import excel_operation, util
import os
"""
getExcelParser
""" 
def parse( fileName, keyMap: dict, productType: list) -> excel_operation.ExcelParser:
    excel = openpyxl.load_workbook(fileName)
    
    result = {}
    for sheet in excel.worksheets:
        sheetName = sheet.title
        if sheetName != "Summary":
            r = excel_operation.ExcelParser( typeRow = 7, originalProductTypeList = productType)
            r.createMap(sheet)
            items = r.parse(fileName, sheetName, 8)
            result[sheetName] = items
    return result

def output(outputFileName, title: list, result: dict ):
    wb  = openpyxl.Workbook()
    sheet = wb.active
    
    sheet.cell(row = 1, column = 1).value = "sheet"
    c = 2
    for t in title:
        sheet.cell(row=1, column=c).value = t
        c += 1

    r = 2
    for key in result.keys():
        items: list[excel_operation.Item]= result[key]
        for item in items:
            sheet.cell(row = r, column = 1).value = key
            c = 2
            for p in item.priority:
                sheet.cell(row = r, column = c).value = p
                c += 1
            r += 1
    wb.save(outputFileName)


def ptcrb_parser():
    fileName = "E:\\04.SourceCode\\Python\\03.ProductManagement\\02.Development\\Python\\02.pms_backend\\pms_backend\\excel_common\\input_test\\gcf.xlsx"
    keyMap = {}
    title = ["SPEC","TC","Band Name","CONDITION","SCS","BW","Description","Band Applicability","PTCRB","GCF","Result","Date","OUT","Platform","Lab","Note"]
    
    result = parse(fileName, keyMap, title)

    outputFileName = util.OUTPUT_FOLDER+"gcf_item.xlsx"

    output(outputFileName, title, result)
    return

if __name__ == '__main__':
    ptcrb_parser()