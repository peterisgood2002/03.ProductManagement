from itertools import product
import string
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import re


KEYINFO = ["Id", "Description", "Status", "ChapterId", "Chapter", "SectionId", "Section", "DocLocation"]
NOCOLUMNINX = -1

"""
   class ExcelParser
"""
class ExcelParser:
    def  __init__(self):
        self.initEmpty()
    
    def  __init__(self, keyRow:int, originalKeyMap: dict, typeRow:int, originalProductTypeList: list):
        self.initEmpty()
        self.initialize(keyRow, originalKeyMap, typeRow, originalProductTypeList)
        
    """
       Initialize method
    """    
    def initEmpty( self ):
        self.keyMap = {}
        self.typeMap = {}
    
    def initialize( self, keyRow:int, originalKeyMap: dict, typeRow:int, originalProductTypeList: list):
        self._originalKeyMap = originalKeyMap.copy()  
        self.keyRow = keyRow
        self._originalProductTypeList = originalProductTypeList.copy()
        self.typeRow = typeRow
            
    """
       END Initialize method
    """ 
    def createMap( self, sheet: Worksheet):
        self.setKeyMap(sheet, self.keyRow, self._originalKeyMap)
        self.setTypeMap(sheet, self.typeRow, self._originalProductTypeList)
    
    """
      the input of These Group function is <KeyInfo, ExcelColumn> and it will create < KeyInfo, ExcelIndex>
    """    
    def setKeyMap( self, sheet, row, keyMap: dict):
        cellValueList = []
        for k in KEYINFO:
            val = keyMap.get(k)
            
            if val == None or val == "":
                print("Please note the key = " + k + " is NULL or empty string." )
                val = NOCOLUMNINX
            cellValueList.append( val )
        index = self.__getColumnsIdxBasedOnList( sheet, row, cellValueList)   
                    
        self.__setMap( sheet, row, index, KEYINFO, self.keyMap)
       
    def setTypeMap( self, sheet, row, productTypeList: list):
        index = self.__getColumnsIdxBasedOnList( sheet, row, productTypeList)  
        self.__setMap( sheet, row, index, productTypeList, self.typeMap)
    
    def __setMap(self, sheet, row, index: list, cellValueList: list, dataMap: map):
        i = 0
        for key in cellValueList:
            dataMap[key] =index[i]
            i = i + 1
                
    def __getColumnsIdxBasedOnList(self, sheet, row, cellValueList: list) -> list:
        if( isinstance(cellValueList, list) != True ):
            raise Exception("Note that cellValueList is a list")

        result = cellValueList.copy()
        
        idxMap = ExcelParser.__getIdxMap(sheet, row, cellValueList)
       
        
        for i in range( 0, len(cellValueList) ):
             result[i] = idxMap[ cellValueList[i] ]      
        return result 
    
    @staticmethod
    def __getIdxMap(sheet: Worksheet, row, cellValueList: list ) -> dict: 
        idxMap = {}
        for v in cellValueList:
            if idxMap.get(v) == None:
                idxMap[v] = -1
        
        for col in sheet.iter_cols(min_row=row, max_row=row):
            cell = col[0]
            if cell.value:
                index = idxMap.get(cell.value)
                if index != None and index == -1:
                    idxMap[cell.value] = cell.col_idx - 1
        return idxMap
    
        
    def getKeyIndex(self, key):
        return self.keyMap[key] 
    def getProductTypeList(self):
        return self._originalProductTypeList
    
    """
        The List of keyColumn should follow the sequence: status, capability, id, desc, section.
        The Priority value depends on the List of productColumn
    """
    def parse( self, fileName, sheetName, dataRow) :
        excel = openpyxl.load_workbook(fileName)
        sheet = excel[sheetName]
        
        print("[ExcelParser.parse][BEGIN]: " + fileName +" MAX ROW = " + str(sheet.max_row))
        result = self.parseExcel(fileName, dataRow, sheet)
        print("[ExcelParser.parse][END]: " + fileName)
        
        excel.close()
        return result

    def parseExcel(self, fileName, dataRow, sheet: Worksheet):
        result = []
        print("[ExcelParser.parseExcel][BEGIN]: MAX ROW = " + str(sheet.max_row))
        idx = 0
        for row in sheet.iter_rows(min_row= dataRow): 
            item = Item(fileName, self.getProductTypeList() )
            item.setKeyInfo( row, self.keyMap, self.typeMap)
            result.append(item)

            idx = idx + 1
            if idx % 500 == 0:
                print(" processing: " + str(idx) )
                
        print("[ExcelParser.parseExcel][END]:")
        return result
"""
   END class ExcelParser
"""
   
"""
  Class Items
"""
class Item:
    def __init__(self, fileName, productType: list):
        self.fileName = fileName
        self.keyValue = []
        self.type = productType.copy()
        self.priority = []  
    def __str__(self):
     return self.id
    
    def setKeyInfo( self, row, keyMap: dict, typeMap: dict):
        # set key 
        for k in KEYINFO:
            idx = keyMap[k]
            if( idx == -1 ):
                self.keyValue.append(None)
            else:
                self.keyValue.append( row[idx].value )
        #add priority
        for t in self.type:
            idx = typeMap[t]
            self.addPriority( row[idx].value )
        
    def addPriority( self, priority):
        self.priority.append(priority)
    def addType( self, t):
        self.type.append(t)
        
    def getKeyInfo( self, key):
        for i in range( len(KEYINFO) ):
            if KEYINFO[i] == key:
                return self.keyValue[i]
           
        return None
    
"""
  End Class Items
"""

#
def getColumnsBasedOnList( sheet, row, cellValueList: list):
    if( isinstance(cellValueList, list) != True ):
         raise Exception("Note that cellValueList is a list")

    result = cellValueList.copy()

    for col in sheet[row]:
        if col.value:
           index = cellValueList.index(col.value) if col.value in cellValueList else -1
           if index != -1:
                result[index] = col.column_letter
                
    return result

def getColumnsBasedOnString( sheet, row, cellValue: str):   
    if( isinstance(cellValue, str) != True ):
        raise Exception("Note that cellValueList is a list")
    for col in sheet[row]:
        if col.value and col.value == cellValue:
            return col.column_letter
def getColumnsIdxBasedOnString( sheet, row, cellValue: str):   
    if( isinstance(cellValue, str) != True ):
        raise Exception("Note that cellValueList is a list")
    for col in sheet[row]:
        if col.value and col.value == cellValue:
            return col.col_idx - 1



   

