
from itertools import product
from pickle import TRUE
from typing import Dict
from common import excel_operation, util
from common import file_operation
import openpyxl
import time
import datetime
import os


"""
getExcelParser
""" 
def getExcelParser( fileName, sheetName, keyMap: dict, productType: list) -> excel_operation.ExcelParser:
    excel = openpyxl.load_workbook(fileName)
    sheet = excel[sheetName]
    result = excel_operation.ExcelParser(4, keyMap, 3, productType)
    result.createMap(sheet)
    return result

"""
END getExcelParser
"""
"""
getItemsFromExcel
""" 
#End getExcelIndex Function
#addToMap Function
def addToMap( items: list[ excel_operation.Item ], itemMap: dict ):
    for item in items:
        id = item.getKeyInfo(excel_operation.KEYINFO.Id)
        if id != None:
            status = item.getKeyInfo(excel_operation.KEYINFO.Status)
            if status != "Delete":
                itemMap[id] = item
#End addToMap Function
def getItemsFromExcel( fileName: list, sheetName, excelParser: excel_operation.ExcelParser ):
    result = {}
  
    beginT = str(datetime.datetime.now())
    #fileName = ["D:\\02.Operator\\02.Softbank\\02.Operator Terminal Requirements\\OTR-2022.Q2.Q3\\OTR-20210730_FullPackage\\02_Conformance_sheet\\OTR-MTC-MTC_RF-CONF-RevD01-20210226_E.xlsx"]
    for f in fileName:
        items = excelParser.parse(f, sheetName, 5)
        print("addToMap: " + f)
        addToMap(items, result)
        print("finish: " + f)
    #items.append( getItems(sheet, keyColumn, productColumn) )
    endT = str(datetime.datetime.now())
    print(beginT)
    print(endT)

    return result
"""
End getItemsFromExcel
"""

"""   
outputExcel
"""
def outputExcel( outputFileName, productType: list, itemMap: dict  ):
    wb  = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1).value = "Id"
    sheet.cell(row=1, column=2).value = "Description"
    sheet.cell(row=1, column=3).value = "Section"
    sheet.cell(row=1, column=4).value = "FileName"
    c = 5
    for p in productType:
        sheet.cell(row=1, column=c).value = p
        c += 1

    r = 2
    for item in itemMap.values():
        sheet.cell(row=r, column=1).value = item.getKeyInfo(excel_operation.KEYINFO.Id)
        sheet.cell(row=r, column=2).value = item.getKeyInfo(excel_operation.KEYINFO.Description)
        sheet.cell(row=r, column=3).value = item.getKeyInfo(excel_operation.KEYINFO.Section)
        sheet.cell(row=r, column=4).value = item.fileName
        c = 5
        for p in item.priority:
            sheet.cell(row=r, column=c).value = p
            c += 1
        
        r += 1
   
    wb.save(outputFileName)
"""   
End outputExcel
"""    

def softbank_parser():
    #0. Paramters
    folder="D:\\02.Operator\\02.Softbank\\02.Operator Terminal Requirements\\OTR.for.2023.Q2.2023.Q3_SB.CPE_20221130\\02_Conformance_sheet\\"
    prefix="" #regular expression
    
    #1. Get fileName in the folder
    fileName = file_operation.getFileNameList(folder, prefix)
    #2. Get each items in the files
    sheetName="Conformance Sheet"
    keyColumnName = {
        excel_operation.KEYINFO.Id: "Requirement ID", 
        excel_operation.KEYINFO.Description:  "Function\n(Description)", 
        excel_operation.KEYINFO.Status: "Status", 
        #"ChapterId":"", 
        excel_operation.KEYINFO.Chapter: "Capability", 
        excel_operation.KEYINFO.SectionId: "Reference\n(Section Info.)", 
        excel_operation.KEYINFO.Section: "Reference\n(Section Name)"
    }
    productType = ["M2M", "Manufacturer brand M2M", "Air"]
    #2.1 Get Critical Column
    excelIndex = getExcelParser( fileName[0], sheetName, keyColumnName, productType)
    itemMap = getItemsFromExcel(fileName, sheetName, excelIndex)
    #3. Output to one Excels
    if os.path.isdir(util.OUTPUT_FOLDER) != True:
        os.makedirs(util.OUTPUT_FOLDER)
    
    outputFileName = util.OUTPUT_FOLDER+"softbank_item.xlsx"
    outputExcel(outputFileName, productType, itemMap)

if __name__ == '__main__':
    softbank_parser()